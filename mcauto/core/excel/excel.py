import win32com.client as win32
import openpyxl as oxl
ACCOUNTING_FORMAT = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
PERCENTAGE_FORMAT = '0.00 %'
import string
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule

conv_dict = {num+1: letter for num, letter in enumerate(string.ascii_uppercase)}

redFill = PatternFill(start_color='FFC7CE',
                end_color='FFC7CE',
                fill_type='solid')
red_text = Font(color="000000",name='Consolas')

greenFill = PatternFill(start_color='C6EFCE',
                end_color='C6EFCE',
                fill_type='solid')
green_text = Font(color="000000", name='Consolas')

blankFill = PatternFill(start_color='222222',
                end_color='222222',
                fill_type='solid')

class ExcelUtility:
    @staticmethod
    def autofit_columns(filename):
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        wb = excel.Workbooks.Open(filename)

        for sheet in wb.Sheets:
            sheet.Columns.AutoFit()

        wb.Save()
        excel.Application.Quit()
        return True

    @staticmethod
    def format_triggers(filename, sheetname=None):

        wb = oxl.load_workbook(filename)
        try:
            if sheetname:
                ExcelUtility._format_trigger(wb, sheetname)
            else:
                for sheetname in wb.sheetnames:
                    ExcelUtility._format_trigger(wb, sheetname)
            wb.save(filename)
        finally:
            wb.close()

    @staticmethod
    def _format_trigger(wb, sheetname):


        if sheetname not in wb.sheetnames:
            raise ValueError(f'{sheetname} not in workbook. Available names are {wb.sheetnames}.')

        ws = wb[sheetname]

        end_cols = False
        end_rows = True
        col_ctr = 1
        max_row = 1

        while True:
            if ws.cell(max_row + 1, 1).value is None:
                break
            max_row += 1

        while True:
            row_ctr = 1
            if ws.cell(1, col_ctr).value is None:
                break

            col_header = ws.cell(row_ctr, col_ctr).value

            # Percent format
            if '%' in col_header:
                fmt = PERCENTAGE_FORMAT
                cdn_fmt = True

            # Accounting format
            else:
                fmt = ACCOUNTING_FORMAT
                cdn_fmt = False


            for i in range(2, max_row+1):
                # if no line items: check first
                if ws.cell(i, col_ctr).value is None:
                    pass
                if cdn_fmt:
                    cell = str(conv_dict[col_ctr]) + str(i)
                    ws.conditional_formatting.add(cell, CellIsRule(operator='lessThan', formula=['0'], fill=redFill,
                                                                   font=red_text))
                    ws.conditional_formatting.add(cell, CellIsRule(operator='greaterThan', formula=['0'], fill=greenFill,
                                                                   font=green_text))
                    ws.conditional_formatting.add(cell, CellIsRule(formula=[f'ISBLANK({cell})'], fill=blankFill))
                    #ws.cell(row_ctr, col_ctr).conditional_formatting.add(cell, )
                ws.cell(i, col_ctr).number_format = fmt


            col_ctr += 1


